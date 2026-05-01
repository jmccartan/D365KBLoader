# =====================================================================
# D365 Knowledge Base Loader
# Copyright (c) 2026 John McCartan
# Licensed under the MIT License. See the LICENSE file in the project
# root for the full text.
# =====================================================================

"""Excel run log generator.

Creates a timestamped .xlsx log file for each run with details on every file processed.
"""

import logging
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

# Column definitions: (header, width)
COLUMNS = [
    ("File Name", 30),
    ("Folder Path", 40),
    ("File Size (bytes)", 18),
    ("Has Content", 14),
    ("HTML Saved", 14),
    ("Published to KB", 16),
    ("KB Action", 16),
    ("Article ID", 38),
    ("Error", 50),
]

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
YES_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NO_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SKIP_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


class RunLog:
    """Collects per-file results and writes them to an Excel log file."""

    def __init__(self):
        self.rows: list[dict] = []
        self.run_start = datetime.now(timezone.utc)
        self.pre_counts: dict[str, int] = {}
        self.post_counts: dict[str, int] = {}

    def set_pre_counts(self, counts: dict[str, int]):
        """Set the KB article counts snapshot taken before processing."""
        self.pre_counts = counts

    def set_post_counts(self, counts: dict[str, int]):
        """Set the KB article counts snapshot taken after processing."""
        self.post_counts = counts

    def add_entry(
        self,
        file_name: str,
        folder_path: str = "",
        file_size: int = 0,
        has_content: bool = False,
        html_saved: bool = False,
        published: bool = False,
        kb_action: str = "",
        article_id: str = "",
        error: str = "",
    ):
        """Record the result of processing one file."""
        self.rows.append({
            "file_name": file_name,
            "folder_path": folder_path,
            "file_size": file_size,
            "has_content": has_content,
            "html_saved": html_saved,
            "published": published,
            "kb_action": kb_action,
            "article_id": article_id,
            "error": error,
        })

    def save(self, output_dir: str) -> Path:
        """Write the log to a timestamped .xlsx file and return the path."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Run Log"

        # -- Summary section --
        run_end = datetime.now(timezone.utc)
        ws.append(["D365 Knowledge Base Loader — Run Log"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
        ws["A1"].font = Font(bold=True, size=14)

        ws.append(["Run started:", self.run_start.strftime("%Y-%m-%d %H:%M:%S UTC")])
        ws.append(["Run finished:", run_end.strftime("%Y-%m-%d %H:%M:%S UTC")])
        ws.append(["Total files:", len(self.rows)])
        ws.append([
            "Summary:",
            f"{sum(1 for r in self.rows if r['published'])} published, "
            f"{sum(1 for r in self.rows if r['kb_action'] == 'Updated')} updated, "
            f"{sum(1 for r in self.rows if r['kb_action'] == 'Skipped')} skipped, "
            f"{sum(1 for r in self.rows if r['error'])} errors",
        ])
        ws.append([])  # blank row

        # -- KB Article Counts (Before & After) --
        if self.pre_counts or self.post_counts:
            all_statuses = sorted(
                set(list(self.pre_counts.keys()) + list(self.post_counts.keys())) - {"Total"}
            )

            ws.append(["KB Article Counts"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

            # Header row for counts table
            ws.append(["Status", "Before Run", "After Run", "Change"])
            count_header_row = ws.max_row
            for col_idx in range(1, 5):
                cell = ws.cell(row=count_header_row, column=col_idx)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
                cell.border = THIN_BORDER
            ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 30)

            for status in all_statuses:
                before = self.pre_counts.get(status, 0)
                after = self.post_counts.get(status, 0)
                change = after - before
                change_str = f"+{change}" if change > 0 else str(change)
                ws.append([status, before, after, change_str])
                current_row = ws.max_row
                for col_idx in range(1, 5):
                    ws.cell(row=current_row, column=col_idx).border = THIN_BORDER
                # Highlight non-zero changes
                change_cell = ws.cell(row=current_row, column=4)
                if change > 0:
                    change_cell.fill = YES_FILL
                elif change < 0:
                    change_cell.fill = NO_FILL

            # Total row
            before_total = self.pre_counts.get("Total", 0)
            after_total = self.post_counts.get("Total", 0)
            total_change = after_total - before_total
            total_change_str = f"+{total_change}" if total_change > 0 else str(total_change)
            ws.append(["Total", before_total, after_total, total_change_str])
            total_row = ws.max_row
            for col_idx in range(1, 5):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.font = Font(bold=True)
                cell.border = THIN_BORDER

            ws.append([])  # blank row

        # -- Header row --
        header_row_num = ws.max_row + 1
        headers = [col[0] for col in COLUMNS]
        ws.append(headers)

        for col_idx, (_, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=header_row_num, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            ws.column_dimensions[cell.column_letter].width = width

        # -- Data rows --
        for row_data in self.rows:
            data_row = [
                row_data["file_name"],
                row_data["folder_path"],
                row_data["file_size"],
                "Yes" if row_data["has_content"] else "No",
                "Yes" if row_data["html_saved"] else "No",
                "Yes" if row_data["published"] else ("Skipped" if row_data["kb_action"] == "Skipped" else "No"),
                row_data["kb_action"],
                row_data["article_id"],
                row_data["error"],
            ]
            ws.append(data_row)
            current_row = ws.max_row

            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=current_row, column=col_idx).border = THIN_BORDER

            # Color-code the boolean/status cells
            for col_idx in (4, 5, 6):  # Has Content, HTML Saved, Published
                cell = ws.cell(row=current_row, column=col_idx)
                if cell.value == "Yes":
                    cell.fill = YES_FILL
                elif cell.value == "No":
                    cell.fill = NO_FILL
                elif cell.value == "Skipped":
                    cell.fill = SKIP_FILL

        # Auto-filter on the header row
        ws.auto_filter.ref = f"A{header_row_num}:{chr(64 + len(COLUMNS))}{ws.max_row}"

        # Freeze panes below the header
        ws.freeze_panes = f"A{header_row_num + 1}"

        # Save file
        log_dir = Path(output_dir)
        log_dir.mkdir(parents=True, exist_ok=True)
        timestamp = self.run_start.strftime("%Y%m%d_%H%M%S")
        log_path = log_dir / f"kb_loader_log_{timestamp}.xlsx"
        wb.save(str(log_path))

        logger.info(f"Run log saved: {log_path}")
        return log_path
